import streamlit as st
import pandas as pd
import numpy as np
from io import BytesIO
from datetime import datetime
import re
import warnings
warnings.filterwarnings("ignore")

st.set_page_config(
    page_title="StockFlow Retail",
    page_icon="📦",
    layout="wide",
    initial_sidebar_state="collapsed"
)

# -------------------------------------------------
# CSS
# -------------------------------------------------
st.markdown("""
<style>
    .stApp {
        background: linear-gradient(rgba(15, 23, 42, 0.78), rgba(15, 23, 42, 0.85)),
                    url('https://images.unsplash.com/photo-1556911220-bff31c812dba?ixlib=rb-4.0.3&auto=format&fit=crop&w=1920&q=80');
        background-size: cover;
        background-position: center;
        background-attachment: fixed;
    }
    
    #MainMenu, footer, header {visibility: hidden;}
   
    .login-card {
        background: rgba(255, 255, 255, 0.96);
        border-radius: 20px;
        padding: 2.2rem 2.5rem;
        box-shadow: 0 25px 50px -12px rgba(0, 0, 0, 0.4);
        border-top: 6px solid #dc2626;
        max-width: 440px;
        margin: 0 auto;
        backdrop-filter: blur(8px);
    }
   
    .module-card {
        background: rgba(255, 255, 255, 0.95);
        border-radius: 18px;
        padding: 1.9rem 1.5rem;
        box-shadow: 0 10px 25px -5px rgba(0, 0, 0, 0.25);
        border: 1px solid rgba(255,255,255,0.3);
        transition: all 0.28s ease;
        height: 100%;
        text-align: center;
        backdrop-filter: blur(6px);
    }
    .module-card:hover {
        transform: translateY(-8px);
        box-shadow: 0 25px 35px -5px rgba(220, 38, 38, 0.3);
        border-color: #f87171;
        background: rgba(255, 245, 245, 0.97);
    }
   
    .logo-circle {
        width: 80px;
        height: 80px;
        border-radius: 50%;
        display: flex;
        align-items: center;
        justify-content: center;
        font-size: 1.9rem;
        margin: 0 auto 1.1rem auto;
    }
   
    .transfer { background: #fee2e2; color: #dc2626; }
    .health   { background: #dcfce7; color: #16a34a; }
    .lpo      { background: #ffedd5; color: #ea580c; }
   
    .stButton > button {
        background: #dc2626 !important;
        color: white !important;
        border: none !important;
        border-radius: 10px !important;
        font-weight: 600 !important;
    }
    .stButton > button:hover {
        background: #b91c1c !important;
    }
   
    h1, h2, h3, h4, h5, h6, p, span, div, label {
        color: #f3f4f6 !important;
    }
    
    .login-card h1, .login-card h2, .login-card h3, 
    .login-card h4, .login-card h5, .login-card p,
    .login-card label, .login-card span, .login-card div {
        color: #1f2937 !important;
    }
    
    .module-card h1, .module-card h2, .module-card h3, 
    .module-card h4, .module-card p, .module-card span {
        color: #1f2937 !important;
    }
    
    .stTextInput > div > div > input {
        background-color: #ffffff !important;
        color: #111827 !important;
        border: 1.5px solid #d1d5db !important;
        border-radius: 8px !important;
    }
    
    button[data-baseweb="tab"] {
        color: #e5e7eb !important;
    }
    
    .user-badge {
        background: #fee2e2;
        color: #dc2626;
        padding: 3px 10px;
        border-radius: 20px;
        font-size: 0.8rem;
        font-weight: 600;
    }
    
    .footer-text {
        text-align: center;
        color: #d1d5db !important;
        font-size: 0.95rem;
        margin-top: 2.5rem;
    }
</style>
""", unsafe_allow_html=True)

# -------------------------------------------------
# Auth State
# -------------------------------------------------
if "users" not in st.session_state:
    st.session_state.users = {}

st.session_state.users["Administrator"] = "Jose2024"

if "demo" not in st.session_state.users:
    st.session_state.users["demo"] = "demo123"
if "manager" not in st.session_state.users:
    st.session_state.users["manager"] = "branch123"

if "pending_resets" not in st.session_state:
    st.session_state.pending_resets = []

if "authenticated" not in st.session_state:
    st.session_state.authenticated = False
if "username" not in st.session_state:
    st.session_state.username = None
if "module" not in st.session_state:
    st.session_state.module = None

def is_strong_password(pw: str) -> bool:
    if len(pw) < 6:
        return False
    has_letter = bool(re.search(r"[a-zA-Z]", pw))
    has_number = bool(re.search(r"[0-9]", pw))
    return has_letter and has_number

# -------------------------------------------------
# Login Page
# -------------------------------------------------
def login_page():
    st.markdown("<br><br>", unsafe_allow_html=True)
    col1, col2, col3 = st.columns([1, 1.4, 1])
   
    with col2:
        st.markdown("""
        <div style="text-align:center; margin-bottom: 1.5rem;">
            <div style="font-size: 3.5rem;">🏠📺🍳📦</div>
            <h1 style="margin:0; font-size: 2.5rem; font-weight: 800; color: #ffffff !important;">StockFlow</h1>
            <p style="color: #e5e7eb !important; font-size: 1.1rem;">Retail Inventory Intelligence Platform</p>
        </div>
        """, unsafe_allow_html=True)
       
        tab_login, tab_create, tab_forgot = st.tabs(["Sign In", "Create Account", "Forgot Password"])
        
        with tab_login:
            st.markdown('<div class="login-card">', unsafe_allow_html=True)
            with st.form("login_form"):
                username = st.text_input("Username")
                password = st.text_input("Password", type="password")
                if st.form_submit_button("Sign In", use_container_width=True):
                    if username in st.session_state.users and st.session_state.users[username] == password:
                        st.session_state.authenticated = True
                        st.session_state.username = username
                        st.rerun()
                    else:
                        st.error("Invalid username or password")
            st.markdown('</div>', unsafe_allow_html=True)
        
        with tab_create:
            st.markdown('<div class="login-card">', unsafe_allow_html=True)
            st.markdown("##### Create a new account")
            with st.form("create_form"):
                new_user = st.text_input("Choose Username")
                new_pw = st.text_input("Password", type="password",
                                      help="Must contain letters and numbers (min 6 characters)")
                confirm_pw = st.text_input("Confirm Password", type="password")
                
                if st.form_submit_button("Create Account", use_container_width=True):
                    if not new_user or not new_pw:
                        st.error("Please fill all fields")
                    elif new_user in st.session_state.users:
                        st.error("Username already exists")
                    elif new_pw != confirm_pw:
                        st.error("Passwords do not match")
                    elif not is_strong_password(new_pw):
                        st.error("Password must contain both letters and numbers (min 6 characters)")
                    else:
                        st.session_state.users[new_user] = new_pw
                        st.success("Account created successfully! You can now sign in.")
            st.markdown('</div>', unsafe_allow_html=True)
        
        with tab_forgot:
            st.markdown('<div class="login-card">', unsafe_allow_html=True)
            st.markdown("##### Forgot Password")
            st.caption("A request will be sent to the Admin. No automatic email is sent.")
            
            with st.form("forgot_form"):
                forgot_user = st.text_input("Enter your Username")
                if st.form_submit_button("Request Password Reset", use_container_width=True):
                    if forgot_user not in st.session_state.users:
                        st.error("Username not found")
                    elif forgot_user in st.session_state.pending_resets:
                        st.warning("A reset request for this user is already pending.")
                    else:
                        st.session_state.pending_resets.append(forgot_user)
                        st.success("Request sent to Admin. Please wait for the Admin to reset your password.")
            st.markdown('</div>', unsafe_allow_html=True)
        
        st.markdown("""
        <div class="footer-text">
            Created by Joseph in 2026
        </div>
        """, unsafe_allow_html=True)

# -------------------------------------------------
# Home Page
# -------------------------------------------------
def home_page():
    col_left, col_right = st.columns([5, 1])
    with col_left:
        st.markdown(f"""
        <h2 style="margin-bottom:0; color:#ffffff !important;">Welcome back,
        <span style="color:#fca5a5;">{st.session_state.username}</span></h2>
        <p style="color:#e5e7eb !important;">Select a module or manage your account</p>
        """, unsafe_allow_html=True)
    with col_right:
        if st.button("Logout", use_container_width=True):
            for key in ["authenticated", "username", "module"]:
                if key in st.session_state:
                    del st.session_state[key]
            st.rerun()
   
    st.markdown("---")
   
    c1, c2, c3 = st.columns(3, gap="large")
   
    with c1:
        st.markdown("""
        <div class="module-card">
            <div class="logo-circle transfer">🚚📦🚚</div>
            <h3 style="margin:0.5rem 0;">Transfer Hub</h3>
            <p style="color:#6b7280; font-size:0.95rem;">
                Inter-branch stock transfers<br>
                Move excess to high-demand branches
            </p>
        </div>
        """, unsafe_allow_html=True)
        st.write("")
        if st.button("Open Transfer Hub →", key="btn_transfer", use_container_width=True):
            st.session_state.module = "transfer"
            st.rerun()
   
    with c2:
        st.markdown("""
        <div class="module-card">
            <div class="logo-circle health">❤️</div>
            <h3 style="margin:0.5rem 0;">Stock Health</h3>
            <p style="color:#6b7280; font-size:0.95rem;">
                Overstock • Dead stock<br>
                Aged inventory & Stockouts
            </p>
        </div>
        """, unsafe_allow_html=True)
        st.write("")
        if st.button("Open Stock Health →", key="btn_health", use_container_width=True):
            st.session_state.module = "health"
            st.rerun()
   
    with c3:
        st.markdown("""
        <div class="module-card">
            <div class="logo-circle lpo">📄📋</div>
            <h3 style="margin:0.5rem 0;">Smart LPO</h3>
            <p style="color:#6b7280; font-size:0.95rem;">
                Automated Local Purchase Orders<br>
                for all branches
            </p>
        </div>
        """, unsafe_allow_html=True)
        st.write("")
        if st.button("Open Smart LPO →", key="btn_lpo", use_container_width=True):
            st.session_state.module = "lpo"
            st.rerun()
   
    st.markdown("<br>", unsafe_allow_html=True)
   
    st.markdown("### Account & Users")
    left, center, right = st.columns([1, 2.4, 1])
   
    with center:
        tab1, tab2, tab3 = st.tabs(["👤 My Account", "👥 Users", "🔑 Admin Panel"])
       
        with tab1:
            st.markdown("#### Update Username or Password")
            with st.form("update_account"):
                st.write(f"Current username: **{st.session_state.username}**")
                
                new_username = st.text_input("New Username (leave blank to keep current)")
                current_pw = st.text_input("Current Password", type="password")
                new_pw = st.text_input("New Password (leave blank to keep current)", type="password")
                confirm_pw = st.text_input("Confirm New Password", type="password")
                
                if st.form_submit_button("Save Changes", use_container_width=True):
                    user = st.session_state.username
                    
                    if st.session_state.users.get(user) != current_pw:
                        st.error("Current password is incorrect")
                    else:
                        if new_pw:
                            if new_pw != confirm_pw:
                                st.error("New passwords do not match")
                            elif not is_strong_password(new_pw):
                                st.error("Password must contain letters and numbers (min 6 characters)")
                            else:
                                st.session_state.users[user] = new_pw
                                st.success("Password updated successfully!")
                        
                        if new_username and new_username != user:
                            if new_username in st.session_state.users:
                                st.error("That username is already taken")
                            else:
                                st.session_state.users[new_username] = st.session_state.users.pop(user)
                                st.session_state.username = new_username
                                st.success(f"Username changed to **{new_username}**")
                                st.rerun()
        
        with tab2:
            st.markdown("#### Registered Users")
            for user in list(st.session_state.users.keys()):
                if user == st.session_state.username:
                    st.markdown(f"""
                    <div style="padding:10px 15px; background:#fee2e2; border-radius:10px; margin-bottom:8px; color:#1f2937 !important;">
                        <span class="user-badge">YOU</span> &nbsp; <b>{user}</b>
                    </div>
                    """, unsafe_allow_html=True)
                else:
                    st.markdown(f"""
                    <div style="padding:10px 15px; background:#f3f4f6; border-radius:10px; margin-bottom:8px; color:#1f2937 !important;">
                        <b>{user}</b>
                    </div>
                    """, unsafe_allow_html=True)
        
        with tab3:
            if st.session_state.username != "Administrator":
                st.warning("Only the Administrator can access this panel.")
            else:
                st.markdown("#### Admin Panel")
                
                st.markdown("##### Pending Password Reset Requests")
                if not st.session_state.pending_resets:
                    st.info("No pending reset requests.")
                else:
                    for user in st.session_state.pending_resets[:]:
                        col_a, col_b = st.columns([3, 1])
                        with col_a:
                            st.write(f"**{user}** requested a password reset")
                        with col_b:
                            if st.button("Clear", key=f"clear_{user}"):
                                st.session_state.pending_resets.remove(user)
                                st.rerun()
                
                st.markdown("---")
                
                st.markdown("##### Reset Any User Password")
                with st.form("admin_reset_form"):
                    target_user = st.selectbox("Select user", options=list(st.session_state.users.keys()))
                    new_password = st.text_input("New Password", type="password")
                    confirm_new = st.text_input("Confirm New Password", type="password")
                    
                    if st.form_submit_button("Reset Password", use_container_width=True):
                        if not new_password:
                            st.error("Please enter a new password")
                        elif new_password != confirm_new:
                            st.error("Passwords do not match")
                        elif not is_strong_password(new_password):
                            st.error("Password must contain letters and numbers (min 6 characters)")
                        else:
                            st.session_state.users[target_user] = new_password
                            if target_user in st.session_state.pending_resets:
                                st.session_state.pending_resets.remove(target_user)
                            st.success(f"Password for **{target_user}** has been reset!")
                
                st.markdown("---")
                
                st.markdown("##### Remove User")
                with st.form("delete_user_form"):
                    users_to_delete = [u for u in st.session_state.users.keys() if u != "Administrator"]
                    if not users_to_delete:
                        st.info("No other users to delete.")
                    else:
                        user_to_delete = st.selectbox("Select user to remove", options=users_to_delete)
                        if st.form_submit_button("Delete User", use_container_width=True):
                            del st.session_state.users[user_to_delete]
                            if user_to_delete in st.session_state.pending_resets:
                                st.session_state.pending_resets.remove(user_to_delete)
                            st.success(f"User **{user_to_delete}** has been removed.")
                            st.rerun()

# -------------------------------------------------
# Shared Data Loader
# -------------------------------------------------
def _is_mda_sda(group_value):
    """Identify MDA/SDA groups, including Primary/Secondary variants."""
    if pd.isna(group_value):
        return False

    text = re.sub(r"\s+", " ", str(group_value).strip().lower())
    return bool(re.search(r"\b(?:mda|sda)\b", text))


def _normalise_month_name(value):
    """Return a normalized 3-letter month label where possible."""
    text = str(value).strip().upper()
    return text[:3]


def _month_columns(df):
    """Return detected month columns in calendar order."""
    month_order = [
        "JAN", "FEB", "MAR", "APR", "MAY", "JUN",
        "JUL", "AUG", "SEP", "OCT", "NOV", "DEC"
    ]

    found = []
    for month in month_order:
        col = next(
            (c for c in df.columns
             if str(c).strip().upper()[:3] == month),
            None
        )
        if col is not None:
            found.append((month, col))
    return found


def _is_backstore_sales_layout(df):
    """
    The back-store sales sheet is branch-row based:

        Branch | Brand | ItemGroup | ModelNo | JAN ... AUG | Total sales

    This check prevents JAN/FEB/MAR/... from ever being interpreted as
    branch names.
    """
    cols = [str(c).strip().lower().replace(" ", "") for c in df.columns]

    has_branch = any(c in {"branch", "showroom"} for c in cols)
    has_brand = any("brand" in c for c in cols)
    has_group = any("itemgroup" in c or c == "group" or "category" in c for c in cols)
    has_model = any("modelno" in c or c == "model" or "sku" in c for c in cols)
    months = _month_columns(df)

    return (
        has_branch
        and has_brand
        and has_group
        and has_model
        and len(months) >= 4
    )


def _find_backstore_sales_sheet(xlsx):
    """Find the sheet using the branch-row/month-column layout."""
    # First prefer explicitly named back-store sales sheets.
    explicit = [
        s for s in xlsx.sheet_names
        if (
            ("back" in s.lower() and "sale" in s.lower())
            or "backstore" in s.lower()
            or "back store" in s.lower()
        )
    ]

    checked = []

    for s in explicit + [s for s in xlsx.sheet_names if s not in explicit]:
        if s in checked:
            continue
        checked.append(s)

        try:
            preview = pd.read_excel(xlsx, sheet_name=s, nrows=10)
        except Exception:
            continue

        if _is_backstore_sales_layout(preview):
            return s

    return None


def _find_standard_sales_sheet(xlsx):
    """
    Find the original StockFlow sales layout:
        Brand | Group | Model | BRANCH1 | BRANCH2 | ...
    Do NOT select the branch-row/month-column back-store sheet.
    """
    for s in xlsx.sheet_names:
        try:
            preview = pd.read_excel(xlsx, sheet_name=s, nrows=5)
        except Exception:
            continue

        if _is_backstore_sales_layout(preview):
            continue

        cols_lower = [str(c).strip().lower() for c in preview.columns]

        has_brand = any("brand" in c for c in cols_lower)
        has_group = any("group" in c or "category" in c for c in cols_lower)
        has_model = any("model" in c or "sku" in c for c in cols_lower)

        # Exclude a pure stock sheet.
        if not (has_brand and has_model):
            continue

        non_id_value_cols = []
        for c, original in zip(cols_lower, preview.columns):
            if (
                "brand" not in c
                and "group" not in c
                and "category" not in c
                and "model" not in c
                and "sku" not in c
                and "total" not in c
                and not c.startswith("unnamed")
            ):
                non_id_value_cols.append(original)

        if non_id_value_cols:
            return s

    return None


def _prepare_backstore_sales(xlsx):
    """
    Read the branch-row/month-column back-store sales sheet.

    Expected:
        Branch | Brand | ItemGroup | ModelNo |
        JAN | FEB | MAR | APR | MAY | JUN | JUL | AUG | Total sales

    Branch is ALWAYS taken from the Branch column.
    Month columns are ONLY used as sales history and are never melted
    into a Branch field.
    """
    sheet = _find_backstore_sales_sheet(xlsx)

    if sheet is None:
        return None

    raw = pd.read_excel(xlsx, sheet_name=sheet).copy()

    branch_col = next(
        (
            c for c in raw.columns
            if str(c).strip().lower() in {"branch", "showroom"}
        ),
        None
    )
    brand_col = next(
        (c for c in raw.columns if "brand" in str(c).lower()),
        None
    )
    group_col = next(
        (
            c for c in raw.columns
            if "itemgroup" in str(c).lower()
            or str(c).strip().lower() == "group"
            or "category" in str(c).lower()
        ),
        None
    )
    model_col = next(
        (
            c for c in raw.columns
            if "modelno" in str(c).lower()
            or "model no" in str(c).lower()
            or str(c).strip().lower() == "model"
            or "sku" in str(c).lower()
        ),
        None
    )
    total_col = next(
        (
            c for c in raw.columns
            if "total" in str(c).lower() and "sale" in str(c).lower()
        ),
        None
    )

    month_cols = _month_columns(raw)

    if not all([branch_col, brand_col, group_col, model_col]) or len(month_cols) < 4:
        return None

    result = pd.DataFrame({
        "Branch": raw[branch_col].astype(str).str.strip(),
        "Brand": raw[brand_col].astype(str).str.strip(),
        "ItemGroup1": raw[group_col].astype(str).str.strip(),
        "Model No": raw[model_col].astype(str).str.strip()
    })

    detected_month_names = []
    for month, col in month_cols:
        result[month] = pd.to_numeric(
            raw[col], errors="coerce"
        ).fillna(0)
        detected_month_names.append(month)

    # Prefer the workbook's Total sales, but calculate it from months if
    # it is missing or unreliable.
    calculated_total = result[detected_month_names].sum(axis=1)

    if total_col is not None:
        reported_total = pd.to_numeric(
            raw[total_col], errors="coerce"
        ).fillna(0)

        # If Total sales is blank/zero while monthly values exist, use the
        # calculated total. Otherwise preserve the supplied total.
        result["Backstore_Total_Sold"] = np.where(
            reported_total.abs() > 0,
            reported_total,
            calculated_total
        )
    else:
        result["Backstore_Total_Sold"] = calculated_total

    # Average monthly sales is based ONLY on the detected Jan-Aug-style
    # monthly columns, not on branch names.
    result["Backstore_Monthly_Avg"] = (
        result[detected_month_names].sum(axis=1) / len(detected_month_names)
    ).round(2)

    # Find ANY 4 consecutive calendar months with >=1 sale.
    # With Jan-Aug data this checks:
    # Jan-Apr, Feb-May, Mar-Jun, Apr-Jul, May-Aug.
    result["Backstore_4M_Consecutive"] = False
    result["Backstore_Consecutive_Sales"] = ""

    month_values = detected_month_names

    for row_idx, row in result.iterrows():
        sales = [float(row[m]) for m in month_values]
        found_run = False

        for i in range(len(sales) - 3):
            four = sales[i:i + 4]
            if all(v >= 1 for v in four):
                months = month_values[i:i + 4]
                result.at[row_idx, "Backstore_4M_Consecutive"] = True
                result.at[row_idx, "Backstore_Consecutive_Sales"] = (
                    " | ".join(
                        f"{m}:{int(v)}"
                        for m, v in zip(months, four)
                    )
                )
                found_run = True
                break

        if not found_run:
            result.at[row_idx, "Backstore_Consecutive_Sales"] = ""

    result["Product_Key"] = (
        result["Brand"].astype(str).str.strip()
        + " | "
        + result["ItemGroup1"].astype(str).str.strip()
        + " | "
        + result["Model No"].astype(str).str.strip()
    )

    # Keep the month columns in the normalized dataset for transparent LPO
    # output. They remain sales history columns, never branch identifiers.
    return result

def load_and_prepare(uploaded, include_backstore_sales=False):
    xlsx = pd.ExcelFile(uploaded)

    sales_sheet = _find_standard_sales_sheet(xlsx)
    stock_sheet = next(
        (s for s in xlsx.sheet_names if "stock" in s.lower()),
        None
    )

    if not sales_sheet or not stock_sheet:
        return None, "Could not find both Sales and Stocks sheets."

    df_sales = pd.read_excel(xlsx, sheet_name=sales_sheet)
    df_stock = pd.read_excel(xlsx, sheet_name=stock_sheet)

    def make_key(df):
        brand = next((c for c in df.columns if "brand" in str(c).lower()), None)
        group = next((c for c in df.columns if "group" in str(c).lower()), None)
        model = next((c for c in df.columns if "model" in str(c).lower()), None)

        parts = []
        if brand:
            parts.append(df[brand].astype(str).str.strip())
        if group:
            parts.append(df[group].astype(str).str.strip())
        if model:
            parts.append(df[model].astype(str).str.strip())

        if not parts:
            raise ValueError("Could not identify Brand, Group or Model columns.")

        key = parts[0]
        for p in parts[1:]:
            key = key + " | " + p

        return key, brand, group, model

    df_s = df_sales.copy()
    df_s["Product_Key"], brand_col, group_col, model_col = make_key(df_s)

    fixed = ["Product_Key"]
    if brand_col:
        fixed.append(brand_col)
    if group_col:
        fixed.append(group_col)
    if model_col:
        fixed.append(model_col)

    branches = [
        c for c in df_s.columns
        if c not in fixed
        and "total" not in str(c).lower()
        and not str(c).startswith("Unnamed")
    ]

    sales_long = df_s.melt(
        id_vars=fixed,
        value_vars=branches,
        var_name="Branch",
        value_name="Qty_Sold"
    )

    sales_long["Qty_Sold"] = pd.to_numeric(
        sales_long["Qty_Sold"], errors="coerce"
    ).fillna(0)

    sales_long["Branch"] = sales_long["Branch"].astype(str).str.strip()

    sales_agg = sales_long.groupby(
        ["Product_Key", "Branch"], as_index=False
    ).agg(
        Total_Sold=("Qty_Sold", "sum"),
        **{
            c: (c, "first")
            for c in [brand_col, group_col, model_col]
            if c
        }
    )

    # Existing StockFlow main-sales rule: 8-month cumulative average.
    sales_agg["Monthly_Avg"] = (
        sales_agg["Total_Sold"] / 8
    ).round(2)

    df_st = df_stock.copy()
    df_st["Product_Key"], _, _, _ = make_key(df_st)

    fixed2 = ["Product_Key"]

    branches2 = [
        c for c in df_st.columns
        if c not in fixed2
        and "total" not in str(c).lower()
        and not str(c).startswith("Unnamed")
    ]

    stock_long = df_st.melt(
        id_vars=fixed2,
        value_vars=branches2,
        var_name="Branch",
        value_name="Current_Stock"
    )

    stock_long["Current_Stock"] = pd.to_numeric(
        stock_long["Current_Stock"], errors="coerce"
    ).fillna(0)

    stock_long["Branch"] = stock_long["Branch"].astype(str).str.strip()

    stock_agg = stock_long.groupby(
        ["Product_Key", "Branch"], as_index=False
    )["Current_Stock"].sum()

    df = pd.merge(
        stock_agg,
        sales_agg,
        on=["Product_Key", "Branch"],
        how="outer"
    )

    df["Current_Stock"] = df["Current_Stock"].fillna(0)
    df["Total_Sold"] = df["Total_Sold"].fillna(0)
    df["Monthly_Avg"] = df["Monthly_Avg"].fillna(0)

    if brand_col:
        df = df.rename(columns={brand_col: "Brand"})
    else:
        df["Brand"] = ""

    if group_col:
        df = df.rename(columns={group_col: "ItemGroup1"})
    else:
        df["ItemGroup1"] = ""

    if model_col:
        df = df.rename(columns={model_col: "Model No"})
    else:
        df["Model No"] = ""

    if include_backstore_sales:
        backstore = _prepare_backstore_sales(xlsx)
        return (df, backstore), None

    return df, None

# -------------------------------------------------
# MODULE 1: Transfer Hub
# -------------------------------------------------
def transfer_hub():
    col1, col2 = st.columns([5, 1])
    with col1:
        st.markdown("## 🚚📦 Transfer Hub")
        st.caption("General inter-branch transfer recommendations")
    with col2:
        if st.button("← Back", use_container_width=True):
            st.session_state.module = None
            st.rerun()
   
    st.markdown("---")
   
    uploaded = st.file_uploader("Upload Excel file (Sales + Stocks sheets)", type=["xlsx", "xls"], key="transfer_upload")
   
    if uploaded is None:
        st.info("Upload your Sales + Stocks Excel file.")
        return
   
    df, error = load_and_prepare(uploaded)
    if error:
        st.error(error)
        return
    
    st.success(f"Data loaded • {df['Product_Key'].nunique()} products • {df['Branch'].nunique()} branches")
    
    MIN_KEEP = 1
    transfers = []
    
    for product, group in df.groupby("Product_Key"):
        receivers = group[
            (group["Monthly_Avg"] >= 0.5) &
            (group["Current_Stock"] <= 3)
        ].sort_values("Total_Sold", ascending=False)
        
        if len(receivers) == 0:
            continue
        
        receiver_branches = set(receivers["Branch"])
        
        donors = group[
            (group["Current_Stock"] >= 2) &
            (~group["Branch"].isin(receiver_branches))
        ].sort_values(["Monthly_Avg", "Current_Stock"], ascending=[True, False])
        
        if len(donors) == 0:
            continue
        
        donors = donors.copy()
        donors["Available"] = (donors["Current_Stock"] - MIN_KEEP).clip(lower=0)
        
        for _, rec in receivers.iterrows():
            target = max(2, int(round(rec["Monthly_Avg"] * 2)))
            needed = max(0, target - int(rec["Current_Stock"]))
            if needed <= 0:
                continue
            
            remaining = needed
            
            for idx, don in donors.iterrows():
                if remaining <= 0:
                    break
                if don["Monthly_Avg"] >= rec["Monthly_Avg"] * 0.75:
                    continue
                
                can_give = donors.at[idx, "Available"]
                if can_give <= 0:
                    continue
                
                give = min(remaining, can_give)
                if give > 0:
                    transfers.append({
                        "Receiver Showroom": rec["Branch"],
                        "Brand": rec.get("Brand", ""),
                        "ItemGroup1": rec.get("ItemGroup1", ""),
                        "Model No": rec.get("Model No", ""),
                        "Receiver current stocks": int(rec["Current_Stock"]) if rec["Current_Stock"] > 0 else "-",
                        "Receiver qty sold period": int(rec["Total_Sold"]),
                        "Donor showroom": don["Branch"],
                        "Donor transfer Qty": int(give)
                    })
                    donors.at[idx, "Available"] -= give
                    remaining -= give
    
    report = pd.DataFrame(transfers)
    
    if len(report) == 0:
        st.warning("No transfer available with current rules.")
    else:
        report = report.sort_values(by=["Receiver qty sold period", "Donor transfer Qty"], ascending=[False, False]).reset_index(drop=True)
        st.success(f"**{len(report)}** recommendations • Total units: **{report['Donor transfer Qty'].sum()}**")
        st.dataframe(report, use_container_width=True, height=480)
        
        output = BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            report.to_excel(writer, sheet_name="transfers", index=False)
        st.download_button(
            "⬇️ Download Transfer Report",
            data=output.getvalue(),
            file_name=f"Inter_Branch_Transfers_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )

# -------------------------------------------------
# MODULE 2: Stock Health
# -------------------------------------------------
def stock_health():
    col1, col2 = st.columns([5, 1])
    with col1:
        st.markdown("## ❤️ Stock Health Analyzer")
        st.caption("Overstock • Dead Stock • Stockouts • Slow-moving items")
    with col2:
        if st.button("← Back", use_container_width=True, key="health_back"):
            st.session_state.module = None
            st.rerun()
    
    st.markdown("---")
    
    uploaded = st.file_uploader("Upload Excel file (Sales + Stocks sheets)", type=["xlsx", "xls"], key="health_upload")
    
    if uploaded is None:
        st.info("Upload your Sales + Stocks Excel file to analyse inventory health.")
        return
    
    df, error = load_and_prepare(uploaded)
    if error:
        st.error(error)
        return
    
    df["Status"] = "Healthy"
    df.loc[df["Current_Stock"] <= 0, "Status"] = "Stockout"
    df.loc[(df["Current_Stock"] > 0) & (df["Total_Sold"] == 0), "Status"] = "Dead Stock"
    
    df["Months_Cover"] = np.where(df["Monthly_Avg"] > 0, df["Current_Stock"] / df["Monthly_Avg"], 999)
    df.loc[(df["Status"] == "Healthy") & (df["Months_Cover"] > 4) & (df["Current_Stock"] > 5), "Status"] = "Overstock"
    df.loc[(df["Status"] == "Healthy") & (df["Months_Cover"] > 2.5) & (df["Current_Stock"] > 3), "Status"] = "Slow-moving"
    
    st.markdown("### Key Metrics")
    k1, k2, k3, k4, k5 = st.columns(5)
    k1.metric("Total SKU-Branch", f"{len(df):,}")
    k2.metric("Stockouts", f"{(df['Status']=='Stockout').sum():,}")
    k3.metric("Dead Stock", f"{(df['Status']=='Dead Stock').sum():,}")
    k4.metric("Overstock", f"{(df['Status']=='Overstock').sum():,}")
    k5.metric("Slow-moving", f"{(df['Status']=='Slow-moving').sum():,}")
    
    st.markdown("---")
    
    tab1, tab2, tab3, tab4 = st.tabs(["🔴 Stockouts", "⚫ Dead Stock", "🟠 Overstock", "🟡 Slow-moving"])
    
    display_cols = ["Branch", "Brand", "ItemGroup1", "Model No", "Current_Stock", "Total_Sold", "Monthly_Avg", "Months_Cover", "Status"]
    
    with tab1:
        stockouts = df[df["Status"] == "Stockout"].sort_values("Total_Sold", ascending=False)
        st.write(f"**{len(stockouts)}** items")
        st.dataframe(stockouts[display_cols], use_container_width=True, height=400)
    
    with tab2:
        dead = df[df["Status"] == "Dead Stock"].sort_values("Current_Stock", ascending=False)
        st.write(f"**{len(dead)}** items • These have stock but zero sales")
        st.dataframe(dead[display_cols], use_container_width=True, height=400)
    
    with tab3:
        over = df[df["Status"] == "Overstock"].sort_values("Months_Cover", ascending=False)
        st.write(f"**{len(over)}** items • More than 4 months of cover")
        st.dataframe(over[display_cols], use_container_width=True, height=400)
    
    with tab4:
        slow = df[df["Status"] == "Slow-moving"].sort_values("Months_Cover", ascending=False)
        st.write(f"**{len(slow)}** items")
        st.dataframe(slow[display_cols], use_container_width=True, height=400)
    
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df[display_cols].to_excel(writer, sheet_name="Full_Health", index=False)
        df[df["Status"]=="Stockout"][display_cols].to_excel(writer, sheet_name="Stockouts", index=False)
        df[df["Status"]=="Dead Stock"][display_cols].to_excel(writer, sheet_name="Dead_Stock", index=False)
        df[df["Status"]=="Overstock"][display_cols].to_excel(writer, sheet_name="Overstock", index=False)
        df[df["Status"]=="Slow-moving"][display_cols].to_excel(writer, sheet_name="Slow_moving", index=False)
    
    st.download_button(
        "⬇️ Download Full Stock Health Report",
        data=output.getvalue(),
        file_name=f"Stock_Health_Report_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True
    )

# -------------------------------------------------
# MODULE 3: Smart LPO (REWORKED LOGIC)
# -------------------------------------------------
def smart_lpo():
    col1, col2 = st.columns([5, 1])

    with col1:
        st.markdown("## 📄📋 Smart LPO Generator")
        st.caption(
            "Standard replenishment + back-store virtual-sales ordering"
        )

    with col2:
        if st.button("← Back", use_container_width=True, key="lpo_back"):
            st.session_state.module = None
            st.rerun()

    st.markdown("---")

    uploaded = st.file_uploader(
        "Upload Excel file (Sales + Stocks + optional Back Store Sales)",
        type=["xlsx", "xls"],
        key="lpo_upload"
    )

    if uploaded is None:
        st.info(
            "Upload your Excel workbook containing Sales and Stocks. "
            "For the new back-store logic, also include the Back Store Sales sheet."
        )
        return

    loaded, error = load_and_prepare(
        uploaded,
        include_backstore_sales=True
    )

    if error:
        st.error(error)
        return

    df, backstore_sales = loaded

    if backstore_sales is None:
        st.warning(
            "The Back Store Sales sheet was not detected. "
            "Standard LPO logic will still run, but LDA, Electronics, "
            "Built-in Range and Commercial Appliances will not use virtual-sales ordering."
        )

    st.markdown("### Ordering Settings")
    col_a, col_b, col_c = st.columns(3)

    with col_a:
        target_months = st.slider(
            "Target Months of Cover",
            min_value=0.5,
            max_value=2.0,
            value=1.0,
            step=0.1,
            help="Normal replenishment target based on the existing 8-month average sales."
        )

    with col_b:
        min_sales = st.number_input(
            "Minimum Monthly Sales",
            min_value=0.0,
            max_value=20.0,
            value=1.0,
            step=0.5,
            help="Normal replenishment threshold for the existing sales logic."
        )

    with col_c:
        safety_days = st.number_input(
            "Minimum Days of Cover before ordering",
            min_value=7,
            max_value=45,
            value=20,
            help="Normal replenishment trigger for the existing sales logic."
        )

    # -------------------------------------------------
    # Existing / standard sales logic
    # -------------------------------------------------
    df["Days_of_Cover"] = np.where(
        df["Monthly_Avg"] > 0,
        (df["Current_Stock"] / df["Monthly_Avg"] * 30).round(1),
        999
    )

    df["Target_Stock"] = np.ceil(
        df["Monthly_Avg"] * target_months
    ).astype(int)

    normal_order = np.where(
        (df["Current_Stock"] < df["Target_Stock"]) &
        (df["Days_of_Cover"] < safety_days) &
        (df["Monthly_Avg"] >= min_sales),
        (df["Target_Stock"] - df["Current_Stock"]).clip(lower=0).astype(int),
        0
    )

    df["Suggested_Order"] = normal_order.astype(int)
    df["Order_Reason"] = np.where(
        df["Suggested_Order"] > 0,
        "Standard cumulative-average replenishment",
        ""
    )

    # -------------------------------------------------
    # NEW: Back-store virtual-sales logic
    #
    # These are large items and should not be stock-held heavily.
    # Item groups:
    #   - LDA
    #   - Electronics
    #   - Built-in Range
    #   - Commercial Appliances
    #
    # Rule:
    #   If cumulative back-store sellout > 0 and current stock is 0,
    #   place exactly 1 pc.
    #
    # Maximum order quantity for each such SKU = 1 pc.
    #
    # This treats sales as "virtual sales" rather than building physical
    # stock in the back store.
    # -------------------------------------------------
    BACKSTORE_GROUPS = {
        "LDA",
        "ELECTRONICS",
        "BUILT-IN RANGE",
        "BUILT IN RANGE",
        "BUILTIN RANGE",
        "COMMERCIAL APPLIANCES",
        "COMMERCIAL"
    }

    if backstore_sales is not None and not backstore_sales.empty:
        # Normalize keys and join branch/SKU back-store sellout information.
        b = backstore_sales.copy()

        # Ensure there is exactly one row per Product + Branch.
        b = b.drop_duplicates(
            subset=["Product_Key", "Branch"]
        )

        # Only merge branch-level sales history. Month columns remain
        # sales-history fields and are never turned into branch values.
        month_history_cols = [
            c for c in [
                "JAN", "FEB", "MAR", "APR", "MAY", "JUN",
                "JUL", "AUG", "SEP", "OCT", "NOV", "DEC"
            ]
            if c in b.columns
        ]

        merge_cols = [
            "Product_Key",
            "Branch",
            "Backstore_Total_Sold",
            "Backstore_Monthly_Avg",
            "Backstore_4M_Consecutive",
            "Backstore_Consecutive_Sales"
        ] + month_history_cols

        b = b[merge_cols]

        df = df.merge(
            b,
            on=["Product_Key", "Branch"],
            how="left"
        )

        df["Backstore_Total_Sold"] = df["Backstore_Total_Sold"].fillna(0)
        df["Backstore_Monthly_Avg"] = df["Backstore_Monthly_Avg"].fillna(0)
        for month in [
            "JAN", "FEB", "MAR", "APR", "MAY", "JUN",
            "JUL", "AUG", "SEP", "OCT", "NOV", "DEC"
        ]:
            if month in df.columns:
                df[month] = df[month].fillna(0)
        df["Backstore_4M_Consecutive"] = (
            df["Backstore_4M_Consecutive"]
            .fillna(False)
            .astype(bool)
        )
        df["Backstore_Consecutive_Sales"] = (
            df["Backstore_Consecutive_Sales"]
            .fillna("")
        )

        df["Is_Backstore_Large_Item"] = (
            df["ItemGroup1"]
            .astype(str)
            .str.strip()
            .str.upper()
            .isin(BACKSTORE_GROUPS)
        )

        # For large back-store items, only the absence of physical stock
        # triggers a virtual replenishment order.
        # IMPORTANT:
        # Branch comes from the Back Store Sales "Branch" column.
        # JAN-AUG are sales-history columns only.
        backstore_virtual_order = np.where(
            df["Is_Backstore_Large_Item"] &
            (df["Backstore_Total_Sold"] > 0) &
            (df["Current_Stock"] <= 0),
            1,
            0
        )

        # Never allow this virtual rule to exceed 1 unit.
        backstore_virtual_order = np.minimum(
            backstore_virtual_order.astype(int),
            1
        )

        # Apply the new virtual-sales order.
        df["Suggested_Order"] = np.maximum(
            df["Suggested_Order"].astype(int),
            backstore_virtual_order
        )

        df["Order_Reason"] = np.select(
            [
                df["Is_Backstore_Large_Item"] &
                (backstore_virtual_order == 1) &
                (normal_order > 0),

                df["Is_Backstore_Large_Item"] &
                (backstore_virtual_order == 1)
            ],
            [
                "Back-store virtual sale + standard replenishment",
                "Back-store virtual sale — order 1 pc"
            ],
            default=df["Order_Reason"]
        )

    else:
        df["Backstore_Total_Sold"] = 0
        df["Backstore_4M_Consecutive"] = False
        df["Backstore_Consecutive_Sales"] = ""
        df["Is_Backstore_Large_Item"] = False

    # -------------------------------------------------
    # MDA / SDA / Power Protection
    #
    # These categories intentionally use the ORIGINAL StockFlow
    # cumulative-average replenishment logic only.
    #
    # The 4-month consecutive-sales protection is NOT applied here.
    # -------------------------------------------------

    # -------------------------------------------------
    # Final order tables
    # -------------------------------------------------
    lpo = df[df["Suggested_Order"] > 0].copy()

    lpo = lpo.sort_values(
        ["Branch", "Suggested_Order", "Monthly_Avg"],
        ascending=[True, False, False]
    )

    # Brands requiring the formal LPO document.
    # MIKA remains included, with Samsung, Bosch and Oryx added.
    LPO_BRANDS = {
        "MIKA",
        "SAMSUNG",
        "BOSCH",
        "ORYX"
    }

    lpo_lpo_brands = (
        lpo["Brand"]
        .astype(str)
        .str.strip()
        .str.upper()
        .isin(LPO_BRANDS)
    )

    mika_lpo = lpo[lpo_lpo_brands].copy()

    # Every other brand goes to the Goods issue sheet.
    goods_issue = lpo[~lpo_lpo_brands].copy()

    if len(lpo) == 0:
        st.success(
            "No orders needed right now. Most branches have sufficient stock based on the current rules."
        )
    else:
        st.success(
            f"**{len(lpo)}** order lines generated across "
            f"**{lpo['Branch'].nunique()}** branches"
        )

    st.markdown("### Summary by Branch")

    if len(lpo) > 0:
        branch_summary = (
            lpo.groupby("Branch")
            .agg(
                Items_to_Order=("Suggested_Order", "count"),
                Total_Units=("Suggested_Order", "sum")
            )
            .sort_values("Total_Units", ascending=False)
            .reset_index()
        )
        st.dataframe(
            branch_summary,
            use_container_width=True
        )
    else:
        branch_summary = pd.DataFrame(
            columns=["Branch", "Items_to_Order", "Total_Units"]
        )
        st.info("No branches currently need replenishment.")

    st.markdown("### Detailed Order Suggestions")

    # Branch is the replenishment destination. Month columns are sales
    # history only and are shown after the branch/product information.
    month_export_cols = [
        c for c in [
            "JAN", "FEB", "MAR", "APR", "MAY", "JUN",
            "JUL", "AUG", "SEP", "OCT", "NOV", "DEC"
        ]
        if c in lpo.columns
    ]

    display_cols = [
        "Branch",
        "Brand",
        "ItemGroup1",
        "Model No",
        "Current_Stock",
        "Backstore_Total_Sold",
        "Backstore_Monthly_Avg",
    ] + month_export_cols + [
        "Monthly_Avg",
        "Days_of_Cover",
        "Target_Stock",
        "Backstore_4M_Consecutive",
        "Backstore_Consecutive_Sales",
        "Suggested_Order",
        "Order_Reason"
    ]

    tab_mika, tab_goods = st.tabs([
        f"📄 LPO — MIKA / Samsung / Bosch / Oryx ({len(mika_lpo)})",
        f"📦 Goods issue — Other Brands ({len(goods_issue)})"
    ])

    with tab_mika:
        if len(mika_lpo) > 0:
            st.dataframe(
                mika_lpo[display_cols].style.format({
                    "Monthly_Avg": "{:.2f}",
                    "Backstore_Total_Sold": "{:.0f}",
                    "Days_of_Cover": "{:.1f}"
                }),
                use_container_width=True,
                height=450
            )
        else:
            st.info("No MIKA / Samsung / Bosch / Oryx items require ordering.")

    with tab_goods:
        if len(goods_issue) > 0:
            st.dataframe(
                goods_issue[display_cols].style.format({
                    "Monthly_Avg": "{:.2f}",
                    "Backstore_Total_Sold": "{:.0f}",
                    "Days_of_Cover": "{:.1f}"
                }),
                use_container_width=True,
                height=450
            )
        else:
            st.info("No non-MIKA items require a Goods issue.")

    # -------------------------------------------------
    # Excel output
    #   LPO          = MIKA + Samsung + Bosch + Oryx
    #   Goods issue  = every other brand
    #   Branch_Summary = combined summary
    # -------------------------------------------------
    output = BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        export_cols = [
            "Branch",
            "Brand",
            "ItemGroup1",
            "Model No",
            "Current_Stock",
            "Backstore_Total_Sold",
            "Backstore_Monthly_Avg",
        ] + month_export_cols + [
            "Monthly_Avg",
            "Days_of_Cover",
            "Target_Stock",
            "Backstore_4M_Consecutive",
            "Backstore_Consecutive_Sales",
            "Suggested_Order",
            "Order_Reason"
        ]

        if len(mika_lpo) > 0:
            mika_lpo[export_cols].to_excel(
                writer,
                sheet_name="LPO",
                index=False
            )
        else:
            pd.DataFrame({
                "Message": ["No MIKA / Samsung / Bosch / Oryx orders recommended"]
            }).to_excel(
                writer,
                sheet_name="LPO",
                index=False
            )

        if len(goods_issue) > 0:
            goods_issue[export_cols].to_excel(
                writer,
                sheet_name="Goods issue",
                index=False
            )
        else:
            pd.DataFrame({
                "Message": ["No non-MIKA orders recommended"]
            }).to_excel(
                writer,
                sheet_name="Goods issue",
                index=False
            )

        branch_summary.to_excel(
            writer,
            sheet_name="Branch_Summary",
            index=False
        )

    # Simple workbook formatting.
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        output.seek(0)
        wb = load_workbook(output)

        header_fill = PatternFill("solid", fgColor="1F4E78")
        header_font = Font(color="FFFFFF", bold=True)

        for ws in wb.worksheets:
            ws.freeze_panes = "A2"

            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center"
                )

            for col_idx in range(1, ws.max_column + 1):
                letter = get_column_letter(col_idx)
                max_len = 0

                for cell in ws[letter]:
                    value = "" if cell.value is None else str(cell.value)
                    max_len = max(max_len, len(value))

                ws.column_dimensions[letter].width = min(
                    max(max_len + 2, 10),
                    32
                )

        output = BytesIO()
        wb.save(output)
        output.seek(0)

    except Exception:
        output.seek(0)

    st.download_button(
        "⬇️ Download Smart LPO Excel",
        data=output.getvalue(),
        file_name=f"Smart_LPO_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True
    )

# -------------------------------------------------
# Router
# -------------------------------------------------
if not st.session_state.authenticated:
    login_page()
else:
    if st.session_state.module is None:
        home_page()
    elif st.session_state.module == "transfer":
        transfer_hub()
    elif st.session_state.module == "health":
        stock_health()
    elif st.session_state.module == "lpo":
        smart_lpo()
